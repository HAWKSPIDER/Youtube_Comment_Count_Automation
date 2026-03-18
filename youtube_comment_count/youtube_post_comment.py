import pandas as pd
import time
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Excel file
file_path = r"C:\Users\Ritesh\OneDrive\Desktop\youtube_comment_count\youtube_post_link.xlsx"

# Read links (first column automatically)
df = pd.read_excel(file_path, sheet_name="Links")

links = df.iloc[:, 0].dropna().tolist()

# Chrome settings (background + faster)
options = Options()
options.add_argument("--headless=new")
options.add_argument("--disable-gpu")
options.add_argument("--window-size=1920,1080")
options.add_argument("--disable-blink-features=AutomationControlled")

prefs = {"profile.managed_default_content_settings.images": 2}
options.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(options=options)

wait = WebDriverWait(driver, 10)

comment_counts = []

for link in tqdm(links, desc="Links done", unit="link"):

    driver.get(link)

    # Scroll so comments load
    driver.execute_script("window.scrollTo(0, 800)")

    # Smart wait instead of time.sleep(2)
    try:
        wait.until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="count"]/yt-formatted-string/span[1]'))
        )
    except:
        pass

    # Comment count
    try:
        comment = driver.find_element(By.XPATH, '//*[@id="count"]/yt-formatted-string/span[1]').text
    except:
        try:
            comment = driver.find_element(By.XPATH, '//*[contains(text(),"comment")]').text
        except:
            comment = "0"

    # Convert to integer (strip commas like "1,234" → 1234)
    try:
        comment_counts.append(int(comment.replace(",", "")))
    except:
        comment_counts.append(0)

driver.quit()

# Write comment counts to column B in Links sheet and auto fit columns
wb = load_workbook(file_path)
ws = wb["Links"]

# Write header and data in column B
ws["B1"] = "Comment Count"
ws["A1"].font = Font(bold=True)
ws["B1"].font = Font(bold=True)
for i, count in enumerate(comment_counts, start=2):
    ws.cell(row=i, column=2, value=count)

# Auto fit column A and B
for col in ["A", "B"]:
    max_length = 0
    for cell in ws[col]:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col].width = max_length + 4

wb.save(file_path)
print(" ")
input("Hoo gaya Bacchi! Enter kar...")